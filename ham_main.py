# encoding=utf-8
#-----------------------------------------------------------------------------
# Copyright (c) 2015-2019, Zhiwei YAN (jerod.yan@gmail.com).
# Author: Jerod YAN (jerod.yan@gmail.com)
# Date:   2019-08-27
# Version：0.1
#-----------------------------------------------------------------------------
"""
The codes read all Chinese keywords embraced with double braces first,
which are inside the file 'input.tex', through the regrex expression,
and search them inside a CN-EN key-mapping table with MS excel format to
match the corresponding English keywords. Then, according to the English
keywords, we find the json keywords in a json file. Finally, It joins the
multiple keywords along the json tree and the joined json keywords are
put back into the original latex file in place of those Chinese keywords.

Args:
input.tex: A template document with Latex format, including the marked 
           Chinese characters with double braces that will be subsituted
           by the corresponding English keywords.
input.json: A file with the json format. It is generated by users through
           a website page.
input.xlsx: A Chinese to English Mapping table with MS excel format. It 
           contains at least two columns, Chinese Keywords Column and 
           English Keywords Column.

Returns:
output.tex: A template document with Latex format, including the those 
           substituted and marked English Keywords strings.
"""
import sys
import json
import re

from openpyxl import load_workbook
from flatten_json import flatten

input_tex_file = './input.tex'
input_json_file = './input.json'

input_excel_file = './input.xlsx'
input_excel_sheet_name = '变量名词'
input_excel_cn_col_idx = 'O'
input_excel_en_col_idx = 'N'

output_tex_file = './output.tex'


def hr(msg):
    print(80 * '-')
    print(msg)
    print(80 * '-')

def print_usage():
    print('-'*80)
    print("The program deals with the three files as input: ")
    print("'input.tex', 'input.xlsx', 'input.json'. ")
    print(' ')
    print("The Chinese strings inside the 'input.tex' are substituted by the Enghlish ")
    print("json keywords inside the 'input.json' according to the mapping of input.xlsx.")
    print("You SHOULD rename your input files as 'input.tex', 'input.xlsx', 'input.json'.")
    print(' ')
    print("Note1: In the file 'input.xlsx', you have the sheet named '变量名词'， put the ")
    print("Chinese keywords in the Column 'O' and the English keywords in the Column 'N'.")
    print("Note2: Put all the three input files and the program in the same directory. ")
    print('-'*80)

def main():
    print_usage()
    a= input("Press ENTER key to continue, ... ...")
    # sys.exit(0)

    # Load the dictionary of Chinese to English from the file 'input_excel_file'.
    wb = load_workbook(filename=input_excel_file)
    ws = wb[input_excel_sheet_name]
    cn_cell_list = [cell.value for cell in ws[input_excel_cn_col_idx]]
    en_cell_list = [cell.value for cell in ws[input_excel_en_col_idx]]
    if len(cn_cell_list) == len(en_cell_list):
        print('cn_cell_column Length:', len(cn_cell_list))
        print('en_cell_column Length:', len(en_cell_list))
        cn2en_dict = dict(zip(cn_cell_list, en_cell_list))
    hr("Processing [%s]:OK" %input_excel_file)

    # clean the invalid values
    for (key, value) in (cn2en_dict.items()):
        if value == None or value == '':
            print('Row:', cn_cell_list.index(key), 'Invalid English keywords:', key, value)
    cn2en_dict = {k: v for k, v in cn2en_dict.items() if v != None and v != ''}
    # print('CNtoEN Dict:', cn2en_dict)
    # Load the json keywords from the file 'input_json_file'
    with open(input_json_file, encoding='UTF-8') as json_file:
        data = json.load(json_file)
        joined_keys_list = list(flatten(data, '.').keys())
        joined_keys_list_partA = [x for x in joined_keys_list if not (x.split('.')[-1]).isdigit()]
        joined_keys_list_partB = [x[:-2] for x in joined_keys_list if '0' == (x.split('.')[-1])]
        joined_keys_list = joined_keys_list_partA + joined_keys_list_partB
        # print('Joined Keys:', joined_keys_list_partA, joined_keys_list_partB)
    hr("Processing Json file [%s]: OK." %input_json_file)

    # search and replace the Chinese Keywords in the file 'input_tex_file'
    cn_keywords = [re.findall(r'\{\{.*?\}\}', line) for line in open(input_tex_file, encoding='UTF-8')]
    with open(input_tex_file, 'r', encoding='UTF-8') as file:
        memory_data = file.read()
        for line_id, keys_in_line in enumerate(cn_keywords):
            if keys_in_line:
                # print("lineNo:", line_id, keys_in_line)
                for key in keys_in_line:
                    key = key[2:-2]
                    if key in cn2en_dict and key != 'None':
                        json_keys = [s for s in joined_keys_list if cn2en_dict[key] in s]
                        if json_keys:
                            print("OK   Line(%d): [%s] --> [%s]" %(line_id+1, key, json_keys[0]))
                            old_str = '{{' + key + '}}'
                            new_str = '[(${' + json_keys[0] + '})] '
                            memory_data = memory_data.replace(old_str, new_str)
                        else:
                            print("E001 Line(%d): The English keywords [%s] of [%s] is NOT in the json file,"
                                  " but in the tex file. " %(line_id+1, cn2en_dict[key], key))

                    else:
                        print("E002 Line(%d): The English keywords of [%s] is is NOT in the excel file,"
                            " but in the tex file. " %(line_id+1, key))



    hr("Replacing strings inside the tex file [%s]: OK." %input_tex_file)

    with open(output_tex_file, 'w', encoding='UTF-8') as file:
        file.write(memory_data)

    print('Output the new file:', output_tex_file)
    print('This is DONE!!')
    a= input("Press ENTER key to continue, ... ...")
    sys.exit(0)
    
if __name__ == "__main__":
    main()
    
